import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
from typing import Dict, Any, List

def generate_excel_report(results: Dict[str, Any], output_filename: str):
    """
    Generates an Excel report with highlighted differences.
    """
    try:
        results_dir = "results"
        os.makedirs(results_dir, exist_ok=True)
        
        # Parse timestamp from filename if present or add new one
        # To be consistent with JSON reporter, we'll just prepend/append if needed, 
        # or trust the caller provided a good name.
        # But commonly we append timestamp.
        timestamp = datetime.now().strftime("%H%M%S")
        filename_root, filename_ext = os.path.splitext(os.path.basename(output_filename))
        if not filename_ext: filename_ext = ".xlsx"
        final_filename = f"{filename_root}_{timestamp}{filename_ext}"
        output_path = os.path.join(results_dir, final_filename)

        wb = Workbook()
        
        # 1. Sheet: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Metric", "Value"])
        
        summary_data = results.get("summary", {})
        for k, v in summary_data.items():
            ws_summary.append([k, v])
            
        # Add Errors if any
        if results.get("errors"):
            ws_summary.append(["Errors", str(results["errors"])])
            
        # Add Header Diffs info
        header_diffs = [d for d in results.get("structured_differences", []) if d['type'].startswith('header')]
        if header_diffs:
             ws_summary.append(["Header Issues", "Yes (See differences)"])

        # 2. Sheet: Missing & Additional
        # Missing
        missing_recs = [d for d in results.get("structured_differences", []) if d['type'] == 'missing_records']
        if missing_recs:
            ws_missing = wb.create_sheet("Missing Records")
            
            # Add Description Header
            source_file = missing_recs[0].get("exclusive_to_file", "Unknown")
            ws_missing.append(["Source File:", source_file])
            ws_missing.append(["Status:", "Registros presentes en este archivo pero que FALTAN en el destino de comparación"])
            ws_missing.append([]) # Empty row for spacing
            
            # Check if we have full row data (Enhanced version)
            has_full_rows = any("full_rows" in item for item in missing_recs)
            
            if has_full_rows:
                # Assume schema is consistent, take headers from first available row
                headers_set = False
                for item in missing_recs:
                    rows = item.get("full_rows", [])
                    if not rows: continue
                    
                    if not headers_set:
                        headers = list(rows[0].keys())
                        ws_missing.append(headers)
                        headers_set = True
                    
                    for r in rows:
                        ws_missing.append([r.get(h) for h in headers])
            else:
                # Fallback to IDs only
                ws_missing.append(["File", "Count", "IDs"])
                for item in missing_recs:
                    ids_str = ", ".join(item.get("ids", [])[:100]) # Limit
                    if len(item.get("ids", [])) > 100: ids_str += "..."
                    ws_missing.append([item.get("exclusive_to_file"), item.get("count"), ids_str])
                 
        # Additional
        additional_recs = [d for d in results.get("structured_differences", []) if d['type'] == 'additional_records']
        if additional_recs:
            ws_extra = wb.create_sheet("Additional Records")
            
            # Add Description Header
            source_file = additional_recs[0].get("exclusive_to_file", "Unknown")
            ws_extra.append(["Source File:", source_file])
            ws_extra.append(["Status:", "Registros presentes en este archivo pero AUSENTES en la base de comparación."])
            ws_extra.append([]) # Empty row
             
            has_full_rows = any("full_rows" in item for item in additional_recs)
            
            if has_full_rows:
                headers_set = False
                for item in additional_recs:
                    rows = item.get("full_rows", [])
                    if not rows: continue
                    
                    if not headers_set:
                        headers = list(rows[0].keys())
                        ws_extra.append(headers)
                        headers_set = True
                    
                    for r in rows:
                        ws_extra.append([r.get(h) for h in headers])
            else:
                 ws_extra.append(["File", "Count", "IDs"])
                 for item in additional_recs:
                     ids_str = ", ".join(item.get("ids", [])[:100])
                     if len(item.get("ids", [])) > 100: ids_str += "..."
                     ws_extra.append([item.get("exclusive_to_file"), item.get("count"), ids_str])

        # 3. Sheet: Content Differences (The fun part)
        content_diffs = [d for d in results.get("structured_differences", []) if d['type'] == 'content_mismatch']
        
        if content_diffs:
            ws_diff = wb.create_sheet("Content Differences")
            
            # Styles
            fill_f1 = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light Green
            fill_f2 = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Light Red
            highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow for specific shell? Or Bold text?
            # User asked: Green row for F1, Red row for F2.
            
            # Use headers from the first diff object (assuming consistent schema mostly, or union keys)
            first_diff = content_diffs[0]
            # keys might differ if schema is weird but let's assume common columns
            # We want to display: KEY | SOURCE | ... COLUMNS ...
            
            common_keys = list(first_diff.get("full_row_file1", {}).keys())
            headers = ["KEY", "SOURCE"] + common_keys
            ws_diff.append(headers)
            
            current_row = 2
            
            for diff in content_diffs:
                row_key = str(diff.get("key"))
                row_f1_data = diff.get("full_row_file1", {})
                row_f2_data = diff.get("full_row_file2", {})
                
                # Identify diff columns to highlight
                diff_cols = set(cd['col'] for cd in diff.get("cell_diffs", []))
                
                # Row for File 1
                ws_diff.cell(row=current_row, column=1, value=row_key).fill = fill_f1
                ws_diff.cell(row=current_row, column=2, value="File 1").fill = fill_f1
                
                for idx, col_name in enumerate(common_keys):
                    cell = ws_diff.cell(row=current_row, column=idx+3, value=str(row_f1_data.get(col_name, "")))
                    cell.fill = fill_f1
                    if col_name in diff_cols:
                        cell.font = Font(bold=True, color="006100") # Dark Green text
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                current_row += 1
                
                # Row for File 2
                ws_diff.cell(row=current_row, column=1, value=row_key).fill = fill_f2
                ws_diff.cell(row=current_row, column=2, value="File 2").fill = fill_f2
                
                for idx, col_name in enumerate(common_keys):
                    cell = ws_diff.cell(row=current_row, column=idx+3, value=str(row_f2_data.get(col_name, "")))
                    cell.fill = fill_f2
                    if col_name in diff_cols:
                        cell.font = Font(bold=True, color="9C0006") # Dark Red text
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                current_row += 1
                
                # Optional: Add empty row or spacer? No, user wanted comparison. Adjacent is better.
        
        wb.save(output_path)
        print(f"[INFO] Excel report saved to: {output_path}")
        
    except Exception as e:
        print(f"[ERROR] Failed to generate Excel report: {e}")
